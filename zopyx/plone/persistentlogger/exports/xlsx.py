"""XLSX export."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from .csv import COLUMNS, _cell


def render_xlsx(rows: list[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("XLSX export requires the xlsx extra") from exc
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("events")
    sheet.append(COLUMNS)
    for row in rows:
        sheet.append([_cell(row.get(column)) for column in COLUMNS])
    metadata = workbook.create_sheet("metadata")
    metadata.append(["schema_version", 1])
    metadata.append(["records", len(rows)])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
